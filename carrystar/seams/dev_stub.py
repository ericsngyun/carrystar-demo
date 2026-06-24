"""Claude-owned DEV STUB for the Codex-lane seams (WS-1/2/5).

PURPOSE: run the API + agent loop + UI on the *verified-real* Ross thread
(CS02411883) BEFORE Codex's parsers/store/replay land. Models the real two-act
timeline:

  Beat 1 (order):    order export + ORIGINAL BOL (662/5) + pick slip + Lina's
                     email asking to ADD PO 11667250 (103 ctn heather grey).
  Beat 2 (revision): Lina's follow-up — "will NOT ship PO 11667250, ships in
                     July, disregard the pick slip" — + the REVISED BOL (559/4).

So the agent first flags PO 11667250 as missing, then retracts it when the
revision lands. The tracker's 559/4 is correct.

Data provenance (all REAL unless noted):
  - Tracker seed: the customer's actual Ross sheet (screenshot) — 4 rows / 559,
    incl. their internal WMS tickets 151385-151388 and D-S 06/19@8AM.
  - Order export: data/.../Book6.xlsx (5 POs / 662).
  - Original / revised BOL: the two real CS02411883.docx (662 vs 559).
  - Email bodies: Lina Vincelli (Mark Edwards Apparel) — real, verbatim.

This is an integration affordance (WS-INT), NOT WS-1/2/5. Codex's real modules
register over it via the seam registry. Do not grow product logic here.
"""

from __future__ import annotations

from pathlib import Path

from carrystar.contracts import (
    Beat,
    DocType,
    Mutation,
    MutationStatus,
    MutationType,
    ParsedDoc,
    StatusColor,
    TrackerRow,
)

DATA_DIR = Path(__file__).resolve().parents[2] / "data" / "emails" / "ross-cs02411883"
ROSS_BOL = "CS02411883"

EMAIL1 = (
    "Hi Vanessa, Please see attached load CS02411883 - to be floor loaded. "
    "Pick up will be 06/17 at 9:00 am. We will be adding PO 11667250 to this load. "
    "Please add customer PO 11667250 to 1 side of the carton for import 221777 - "
    "all 103 cartons of color heather grey. Thank you, Lina Vincelli, Mark Edwards Apparel."
)
EMAIL2 = (
    "Hi Vanessa, We will not be shipping PO 11667250 on this load. It will ship in July. "
    "Please disregard the picking slip. I attached the revised BOL. "
    "Thank you, Lina Vincelli, Mark Edwards Apparel."
)


# --- Real Ross tracker seed (from the customer's sheet) ----------------------
# (po, container, etd, import_po, style, ctn, pc, wms_ticket, status)
_SEED = [
    ("11626058", "CAAU4749341", "07 May, 2026", "221415", "76284J-AK", 330, 9900, "151385", StatusColor.PLAIN),
    ("11573709", "CAAU9809171", "12 Mar, 2026", "219094SS", "85739J-AD", 10, 600, "151386", StatusColor.PLAIN),
    ("11573712", "CAAU9809171", "12 Mar, 2026", "219094SS", "85739J-AG", 10, 600, "151387", StatusColor.PLAIN),
    ("11722464", "XYLU8223291", "21 May, 2026", "221915", "85404J-AA", 209, 10002, "151388", StatusColor.RED),
]

# The 5th, rescinded line (heather grey) — present in the ORIGINAL docs only.
_RESCINDED_LINE = ("11667250", "MATU2103718", "27 May, 2026", "221777", "82355J-IU", 103, 3605)


def _initial_rows() -> list[TrackerRow]:
    rows: list[TrackerRow] = []
    for po, cont, etd, imp, style, ctn, pc, wms, status in _SEED:
        rows.append(TrackerRow(
            row_id=f"row-ross-{po}", shipment_id=ROSS_BOL, account="ROSS", bol_number=ROSS_BOL,
            container=cont, date_of_etd=etd, import_po=imp, style=style, customer_po=po,
            ctn_qty=ctn, pallet="FL/LOAD", pc_qty=pc, ds="06/19@8AM", wms_ticket=wms,
            status_color=status,
        ))
    return rows


# ---------------------------------------------------------------------------
# Parsed documents (what WS-1 parsers will eventually emit)
# ---------------------------------------------------------------------------


def _order_export() -> ParsedDoc:
    lines = [(po, cont, etd, imp, style, ctn, pc) for (po, cont, etd, imp, style, ctn, pc, *_ ) in _SEED]
    lines.append(_RESCINDED_LINE)
    rows = []
    for i, (po, cont, etd, imp, style, ctn, pc) in enumerate(lines, start=2):
        rows.append({
            "account": "ROSS", "bol_number": ROSS_BOL, "container": cont, "date_of_etd": etd,
            "import_po": imp, "style": style, "customer_po": po, "ctn_qty": ctn, "pc_qty": pc,
            "_sources": [{"doc_name": "Book6.xlsx", "locator": f"Sheet1 row {i}"}],
        })
    return ParsedDoc(doc_id="ross-order-export", doc_name="Book6.xlsx",
                     doc_type=DocType.ORDER_EXPORT_XLSX, shipment_id=ROSS_BOL, rows=rows, confidence=0.98)


def _bol(revised: bool) -> ParsedDoc:
    # Original: 5 POs / 662. Revised: 4 POs / 559 (PO 11667250 removed).
    lines = [("11722464", 209, "FLOOR LOAD"), ("11573709", 10, ""), ("11573712", 10, ""),
             ("11626058", 330, "")]
    if not revised:
        lines.append(("11667250", 103, ""))
    name = "BOL_CS02411883_revised.docx" if revised else "BOL_CS02411883_original.docx"
    rows = []
    for i, (po, ctn, pallet) in enumerate(lines, start=1):
        rows.append({
            "account": "ROSS", "bol_number": ROSS_BOL, "customer_po": po, "ctn_qty": ctn,
            "pallet": "FL/LOAD" if pallet else "",
            "_sources": [{"doc_name": name, "locator": f"Customer Order Number table, row {i}"}],
        })
    return ParsedDoc(doc_id=f"ross-bol-{'rev' if revised else 'orig'}", doc_name=name,
                     doc_type=DocType.BOL_DOCX, shipment_id=ROSS_BOL, rows=rows, confidence=0.95)


def _pickslip() -> ParsedDoc:
    # Thin text layer — corroboration only.
    return ParsedDoc(doc_id="ross-pickslip", doc_name="Pick Slips - Export.pdf",
                     doc_type=DocType.PICKSLIP_PDF, shipment_id=ROSS_BOL,
                     rows=[{"customer_po": "11667250", "ctn_qty": 103,
                            "_sources": [{"doc_name": "Pick Slips - Export.pdf", "locator": "pick slip"}]}],
                     confidence=0.6)


def _email(num: int) -> ParsedDoc:
    if num == 1:
        return ParsedDoc(doc_id="ross-email1", doc_name="email — Lina Vincelli (add)",
                         doc_type=DocType.EMAIL_BODY, shipment_id=ROSS_BOL,
                         rows=[{"customer_po": "11667250", "ctn_qty": 103, "import_po": "221777",
                                "_sources": [{"doc_name": "email — Lina Vincelli", "locator": "add instruction"}]}],
                         confidence=0.7, notes=EMAIL1)
    # A rescind is a CONTROL signal, not an order line — no rows, so it never
    # looks like a new PO to the engine. The rescind intent rides on beat.rescinds.
    return ParsedDoc(doc_id="ross-email2", doc_name="email — Lina Vincelli (revision)",
                     doc_type=DocType.EMAIL_BODY, shipment_id=ROSS_BOL,
                     rows=[], confidence=0.7, notes=EMAIL2)


# ---------------------------------------------------------------------------
# Dev ParserRegistry — routes a real file path to its ParsedDoc.
# ---------------------------------------------------------------------------


class _DevParsers:
    def parse_path(self, path: Path) -> ParsedDoc:
        n = path.name.lower()
        if n.endswith(".xlsx"):
            return _order_export()
        if "revised" in n:
            return _bol(revised=True)
        if n.endswith(".docx"):
            return _bol(revised=False)
        if n.endswith(".pdf"):
            return _pickslip()
        return _email(1)


# ---------------------------------------------------------------------------
# Dev Store — in-memory tracker + 14-column xlsx mirror.
# ---------------------------------------------------------------------------


class _DevStore:
    def __init__(self) -> None:
        self._rows: list[TrackerRow] = _initial_rows()

    def get_state(self) -> list[TrackerRow]:
        return [r.model_copy(deep=True) for r in self._rows]

    def get_row(self, row_id: str) -> TrackerRow | None:
        return next((r.model_copy(deep=True) for r in self._rows if r.row_id == row_id), None)

    def get_row_by_po(self, shipment_id: str, po: str) -> TrackerRow | None:
        return next((r.model_copy(deep=True) for r in self._rows
                     if r.shipment_id == shipment_id and r.customer_po == po), None)

    def apply_mutation(self, mutation: Mutation) -> TrackerRow:
        if mutation.status not in (MutationStatus.APPROVED, MutationStatus.EDITED):
            raise ValueError(
                f"refusing to apply mutation {mutation.mutation_id} with status "
                f"{mutation.status} — only approved/edited mutations commit"
            )
        if mutation.type == MutationType.ADD_ROW:
            row = mutation.proposed_row or TrackerRow(row_id=mutation.row_id or mutation.mutation_id,
                                                      shipment_id=mutation.shipment_id)
            self._rows.append(row.model_copy(deep=True))
            return row
        if mutation.type == MutationType.REMOVE_ROW:
            for i, r in enumerate(self._rows):
                if r.row_id == mutation.row_id:
                    return self._rows.pop(i)
            raise KeyError(f"row {mutation.row_id} not found for remove_row")
        # update_field
        for r in self._rows:
            if r.row_id == mutation.row_id:
                setattr(r, mutation.field, _coerce(mutation.field, mutation.new_value))
                return r.model_copy(deep=True)
        raise KeyError(f"row {mutation.row_id} not found for update_field")

    def write_mirror_xlsx(self, path: Path) -> Path:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        from carrystar.contracts import TRACKER_COLUMNS

        fills = {"green": "C6EFCE", "blue": "BDD7EE", "red": "FFC7CE"}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tracker"
        ws.append([c.replace("_", " ").upper() for c in TRACKER_COLUMNS])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in self._rows:
            ws.append([getattr(r, col) for col in TRACKER_COLUMNS])
            hexc = fills.get(r.status_color.value)
            if hexc:
                for c in ws[ws.max_row]:
                    c.fill = PatternFill("solid", fgColor=hexc)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path

    def reset(self) -> None:
        self._rows = _initial_rows()


def _coerce(field: str, value: str):
    if field in ("ctn_qty", "pc_qty"):
        try:
            return int(str(value).replace(",", "").strip() or 0)
        except ValueError:
            return 0
    return value


# ---------------------------------------------------------------------------
# Dev ReplaySource — the Ross thread as two ordered beats.
# ---------------------------------------------------------------------------


class _DevReplay:
    def beats(self) -> list[Beat]:
        return [
            Beat(
                beat_id="ross-1-order", kind="order", shipment_id=ROSS_BOL, account="ROSS",
                sender="Lina Vincelli · Mark Edwards Apparel",
                subject="Load CS02411883 — pickup 06/17, adding PO 11667250",
                email_body=EMAIL1,
                attachment_names=["Book6.xlsx", "BOL CS02411883.docx", "Pick Slips - Export.pdf"],
                parsed_docs=[_order_export(), _bol(revised=False), _pickslip(), _email(1)],
            ),
            Beat(
                beat_id="ross-2-revision", kind="revision", shipment_id=ROSS_BOL, account="ROSS",
                sender="Lina Vincelli · Mark Edwards Apparel",
                subject="RE: Load CS02411883 — revised BOL (PO 11667250 pulled)",
                email_body=EMAIL2, rescinds=["11667250"],
                attachment_names=["BOL CS02411883 (revised).docx"],
                parsed_docs=[_bol(revised=True), _email(2)],
            ),
        ]

    def packets(self):
        return []

    def cached_run(self):
        return None


# --- module-level singletons (store is stateful) -----------------------------
_PARSERS = _DevParsers()
_STORE = _DevStore()
_REPLAY = _DevReplay()


def parsers() -> _DevParsers:
    return _PARSERS


def store() -> _DevStore:
    return _STORE


def replay() -> _DevReplay:
    return _REPLAY
