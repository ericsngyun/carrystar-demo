"""SQLite implementation of the WS-2 Store seam."""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

from carrystar.config import CACHE_DIR
from carrystar.contracts import (
    Mutation,
    MutationStatus,
    MutationType,
    StatusColor,
    TRACKER_COLUMNS,
    TrackerRow,
)


def _initial_rows() -> list[TrackerRow]:
    from carrystar.seed import initial_tracker_rows

    return initial_tracker_rows()


def _coerce(field: str | None, value: Any) -> Any:
    if field in {"ctn_qty", "pc_qty"}:
        try:
            return int(str(value).replace(",", "").strip() or 0)
        except (TypeError, ValueError):
            return 0
    return "" if value is None else str(value)


class SQLiteStore:
    def __init__(self, path: str | Path | None = None) -> None:
        self.path = Path(path) if path is not None else CACHE_DIR / "carrystar.sqlite3"
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()
        if not self.get_state():
            self.reset()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS tracker_rows (
                    row_id TEXT PRIMARY KEY,
                    shipment_id TEXT NOT NULL,
                    customer_po TEXT NOT NULL,
                    sort_order INTEGER NOT NULL,
                    data TEXT NOT NULL
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_tracker_shipment_po ON tracker_rows(shipment_id, customer_po)")

    def _insert_row(self, conn: sqlite3.Connection, row: TrackerRow, sort_order: int) -> None:
        conn.execute(
            """
            INSERT OR REPLACE INTO tracker_rows(row_id, shipment_id, customer_po, sort_order, data)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                row.row_id,
                row.shipment_id,
                row.customer_po,
                sort_order,
                row.model_dump_json(),
            ),
        )

    def get_state(self) -> list[TrackerRow]:
        with self._connect() as conn:
            records = conn.execute("SELECT data FROM tracker_rows ORDER BY sort_order, row_id").fetchall()
        return [TrackerRow.model_validate_json(record["data"]) for record in records]

    def get_row(self, row_id: str) -> TrackerRow | None:
        with self._connect() as conn:
            record = conn.execute("SELECT data FROM tracker_rows WHERE row_id = ?", (row_id,)).fetchone()
        return TrackerRow.model_validate_json(record["data"]) if record else None

    def get_row_by_po(self, shipment_id: str, po: str) -> TrackerRow | None:
        with self._connect() as conn:
            record = conn.execute(
                "SELECT data FROM tracker_rows WHERE shipment_id = ? AND customer_po = ? ORDER BY sort_order LIMIT 1",
                (shipment_id, po),
            ).fetchone()
        return TrackerRow.model_validate_json(record["data"]) if record else None

    def apply_mutation(self, mutation: Mutation) -> TrackerRow:
        if mutation.status not in (MutationStatus.APPROVED, MutationStatus.EDITED):
            raise ValueError(
                f"refusing to apply mutation {mutation.mutation_id} with status "
                f"{mutation.status} - only approved/edited mutations commit"
            )

        if mutation.type == MutationType.ADD_ROW:
            row = mutation.proposed_row or TrackerRow(
                row_id=mutation.row_id or mutation.mutation_id,
                shipment_id=mutation.shipment_id,
            )
            with self._connect() as conn:
                next_sort = conn.execute("SELECT COALESCE(MAX(sort_order), -1) + 1 AS n FROM tracker_rows").fetchone()["n"]
                self._insert_row(conn, row, int(next_sort))
            return row.model_copy(deep=True)

        if mutation.type == MutationType.REMOVE_ROW:
            if not mutation.row_id:
                raise ValueError("remove_row mutation requires row_id")
            row = self.get_row(mutation.row_id)
            if row is None:
                raise KeyError(f"row {mutation.row_id} not found for remove_row")
            with self._connect() as conn:
                conn.execute("DELETE FROM tracker_rows WHERE row_id = ?", (mutation.row_id,))
            return row

        if not mutation.row_id or not mutation.field:
            raise ValueError("update_field mutation requires row_id and field")
        if mutation.field not in TRACKER_COLUMNS:
            raise ValueError(f"unknown tracker field {mutation.field}")
        row = self.get_row(mutation.row_id)
        if row is None:
            raise KeyError(f"row {mutation.row_id} not found for update_field")
        setattr(row, mutation.field, _coerce(mutation.field, mutation.new_value))
        with self._connect() as conn:
            current = conn.execute("SELECT sort_order FROM tracker_rows WHERE row_id = ?", (row.row_id,)).fetchone()
            self._insert_row(conn, row, int(current["sort_order"]))
        return row.model_copy(deep=True)

    def write_mirror_xlsx(self, path: Path) -> Path:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        fills = {
            StatusColor.GREEN.value: "C6EFCE",
            StatusColor.BLUE.value: "BDD7EE",
            StatusColor.RED.value: "FFC7CE",
        }
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tracker"
        ws.append([c.replace("_", " ").upper() for c in TRACKER_COLUMNS])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in self.get_state():
            ws.append([getattr(row, col) for col in TRACKER_COLUMNS])
            fill = fills.get(row.status_color.value)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=fill)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path

    def reset(self) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM tracker_rows")
            for sort_order, row in enumerate(_initial_rows()):
                self._insert_row(conn, row, sort_order)


_STORE: SQLiteStore | None = None


def store(path: str | Path | None = None) -> SQLiteStore:
    global _STORE
    if path is not None:
        return SQLiteStore(path)
    if _STORE is None:
        _STORE = SQLiteStore()
    return _STORE
